import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Brand colors
BLUE = "1E6FD9"
GREEN = "22C55E"
DARK = "0F172A"
AMBER = "D97706"
WHITE = "FFFFFF"
OFF_WHITE = "F8FAFC"
LIGHT_GRAY = "F1F5F9"
MID_GRAY = "94A3B8"
TEXT_DARK = "1E293B"
TEXT_GRAY = "64748B"

# Styles
thin_border = Border(
    left=Side(style='thin', color='E2E8F0'),
    right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'),
    bottom=Side(style='thin', color='E2E8F0')
)

header_fill = PatternFill(start_color=DARK, end_color=DARK, fill_type='solid')
green_fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type='solid')
blue_fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type='solid')
amber_fill = PatternFill(start_color=AMBER, end_color=AMBER, fill_type='solid')
light_fill = PatternFill(start_color=OFF_WHITE, end_color=OFF_WHITE, fill_type='solid')
light_gray_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid')
accent_green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type='solid')
accent_blue_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type='solid')
accent_amber_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type='solid')

header_font = Font(name='Arial', bold=True, color=WHITE, size=11)
title_font = Font(name='Arial', bold=True, color=DARK, size=18)
subtitle_font = Font(name='Arial', color=TEXT_GRAY, size=12)
body_font = Font(name='Arial', color=TEXT_DARK, size=11)
stat_font = Font(name='Arial', bold=True, color=BLUE, size=22)
stat_label_font = Font(name='Arial', color=TEXT_GRAY, size=9)

center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def style_range(ws, row_start, row_end, col_start, col_end, font=None, fill=None, border=None, alignment=None):
    for row in ws.iter_rows(min_row=row_start, max_row=row_end, min_col=col_start, max_col=col_end):
        for cell in row:
            if font: cell.font = font
            if fill: cell.fill = fill
            if border: cell.border = border
            if alignment: cell.alignment = alignment

def merge_cells(ws, row_start, row_end, col_start, col_end, value, font=body_font, fill=None, alignment=left_align):
    ws.merge_cells(min_row=row_start, max_row=row_end, min_col=col_start, max_col=col_end)
    cell = ws.cell(row=row_start, column=col_start, value=value)
    cell.font = font
    if fill: cell.fill = fill
    cell.alignment = alignment
    return cell

# ═══════════════════════════════════════
# SHEET 1 — TITLE
# ═══════════════════════════════════════
ws1 = wb.active
ws1.title = "1. Title"
set_col_widths(ws1, [20, 40, 20, 20])

# Dark header row
for c in range(1, 9):
    ws1.cell(row=1, column=c).fill = PatternFill(start_color=DARK, end_color=DARK, fill_type='solid')
ws1.merge_cells('A1:H1')

# Title section
ws1.merge_cells('B3:F3')
ws1.cell(row=3, column=2, value="KlassApp").font = Font(name='Arial', bold=True, color=BLUE, size=42)
ws1.cell(row=3, column=2).alignment = Alignment(horizontal='left', vertical='center')

# Green accent bar (row 4)
for c in range(2, 5):
    ws1.cell(row=4, column=c).fill = green_fill
    ws1.cell(row=4, column=c).font = Font(name='Arial', size=1)

ws1.merge_cells('B6:F6')
ws1.cell(row=6, column=2, value="The School in Every Parent's Pocket").font = Font(name='Arial', color=TEXT_DARK, size=20)
ws1.cell(row=6, column=2).alignment = Alignment(horizontal='left', vertical='center')

ws1.merge_cells('B8:F8')
ws1.cell(row=8, column=2, value="WhatsApp-first school management for Africa").font = Font(name='Arial', color=TEXT_GRAY, size=14)
ws1.cell(row=8, column=2).alignment = Alignment(horizontal='left', vertical='center')

ws1.merge_cells('B12:F12')
ws1.cell(row=12, column=2, value="Confidential — June 2026").font = Font(name='Arial', color=MID_GRAY, size=10)
ws1.cell(row=12, column=2).alignment = Alignment(horizontal='left', vertical='center')

# ═══════════════════════════════════════
# SHEET 2 — PROBLEM
# ═══════════════════════════════════════
ws2 = wb.create_sheet("2. Problem")
set_col_widths(ws2, [4, 50, 4, 4, 16, 16, 14])

# Header — two sections
ws2.merge_cells('A1:D1')
ws2.cell(row=1, column=1, value="THE PROBLEM — The Communication Gap").font = Font(name='Arial', bold=True, color=WHITE, size=16)
ws2.cell(row=1, column=1).fill = header_fill
ws2.cell(row=1, column=1).alignment = Alignment(horizontal='left', vertical='center')
ws2.merge_cells('E1:G1')
ws2.cell(row=1, column=5, value="KEY MARKET STATS").font = Font(name='Arial', bold=True, color=WHITE, size=12)
ws2.cell(row=1, column=5).fill = header_fill
ws2.cell(row=1, column=5).alignment = Alignment(horizontal='center', vertical='center')
style_range(ws2, 1, 1, 1, 7, fill=header_fill)

# Bullets
bullets = [
    "74,000+ schools across Uganda. 20 million students. Zero real-time parent communication.",
    "Parents find out exam results at end of term — weeks after publication.",
    "Schools spend 30 UGX per SMS. 70% delivery rate. One-way. No replies.",
    "Fee collection is chaotic. Absence notifications never reach home.",
    "\"School apps\" require downloads, logins, and updates. Parents don't use them.",
    "Paper circulars are lost, late, and unread.",
]
for i, bullet in enumerate(bullets):
    ws2.merge_cells(f'B{3+i}:C{3+i}')
    ws2.cell(row=3+i, column=2, value=bullet).font = Font(name='Arial', color=TEXT_DARK, size=12)
    ws2.cell(row=3+i, column=2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws2.row_dimensions[3+i].height = 22

# Stat cards section
ws2.merge_cells('E1:G1')
ws2.cell(row=1, column=5, value="KEY MARKET STATS").font = Font(name='Arial', bold=True, color=WHITE, size=12)
ws2.cell(row=1, column=5).alignment = Alignment(horizontal='center', vertical='center')

stats = [
    ("98%", "WhatsApp penetration among smartphone users", GREEN),
    ("$40", "Smartphones now common in Uganda", BLUE),
    ("70%+", "School fees paid via mobile money", AMBER),
]
for i, (num, label, color) in enumerate(stats):
    r = 3 + i * 3
    fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    ws2.merge_cells(f'E{r}:G{r}')
    ws2.cell(row=r, column=5, value=num).font = Font(name='Arial', bold=True, color=WHITE, size=28)
    ws2.cell(row=r, column=5).fill = fill
    ws2.cell(row=r, column=5).alignment = center_align
    ws2.merge_cells(f'E{r+1}:G{r+1}')
    ws2.cell(row=r+1, column=5, value=label).font = Font(name='Arial', color=TEXT_DARK, size=10)
    ws2.cell(row=r+1, column=5).alignment = center_align

# ═══════════════════════════════════════
# SHEET 3 — SOLUTION
# ═══════════════════════════════════════
ws3 = wb.create_sheet("3. Solution")
set_col_widths(ws3, [4, 50, 4, 4, 16, 16, 14])

style_range(ws3, 1, 1, 1, 7, fill=header_fill)
ws3.merge_cells('A1:D1')
ws3.cell(row=1, column=1, value="THE SOLUTION — Meet KlassApp").font = Font(name='Arial', bold=True, color=WHITE, size=16)
ws3.cell(row=1, column=1).fill = header_fill
ws3.cell(row=1, column=1).alignment = Alignment(horizontal='left', vertical='center')
ws3.merge_cells('E1:G1')
ws3.cell(row=1, column=5, value="KEY METRICS").font = Font(name='Arial', bold=True, color=WHITE, size=12)
ws3.cell(row=1, column=5).fill = header_fill
ws3.cell(row=1, column=5).alignment = Alignment(horizontal='center', vertical='center')

ws3.merge_cells('B2:C2')
ws3.cell(row=2, column=2, value="The platform that makes every school a WhatsApp contact.").font = Font(name='Arial', color=TEXT_GRAY, size=12, italic=True)

features = [
    "Parents get grades, fees, and attendance via WhatsApp. No app. No login. Just a message.",
    "Schools get dashboards, automated campaigns, and delivery tracking.",
    "Works on top of existing school systems — no rip-and-replace.",
    "Two-way: parents check balances, report absences, and request results.",
    "Free to receive. WhatsApp is zero-rated on MTN and Airtel in Uganda.",
]
for i, f in enumerate(features):
    ws3.merge_cells(f'B{4+i}:C{4+i}')
    ws3.cell(row=4+i, column=2, value=f).font = Font(name='Arial', color=TEXT_DARK, size=12)
    ws3.cell(row=4+i, column=2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws3.row_dimensions[4+i].height = 22

# Metric cards
metrics = [
    ("0-15 UGX", "per message\n(vs 30 UGX SMS)", GREEN),
    ("95%+", "delivery rate\n(vs 70% SMS)", BLUE),
    ("97%", "gross margin", AMBER),
]
for i, (num, label, color) in enumerate(metrics):
    r = 4 + i * 3
    fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    ws3.merge_cells(f'E{r}:G{r}')
    ws3.cell(row=r, column=5, value=num).font = Font(name='Arial', bold=True, color=WHITE, size=28)
    ws3.cell(row=r, column=5).fill = fill
    ws3.cell(row=r, column=5).alignment = center_align
    ws3.merge_cells(f'E{r+1}:G{r+1}')
    ws3.cell(row=r+1, column=5, value=label).font = Font(name='Arial', color=TEXT_DARK, size=9)
    ws3.cell(row=r+1, column=5).alignment = center_align

# ═══════════════════════════════════════
# SHEET 4 — PRODUCT ARCHITECTURE
# ═══════════════════════════════════════
ws4 = wb.create_sheet("4. Architecture")
set_col_widths(ws4, [5, 25, 25, 25, 25, 10])

style_range(ws4, 1, 1, 1, 6, fill=header_fill)
ws4.merge_cells('A1:F1')
ws4.cell(row=1, column=1, value="PLATFORM — How It Works").font = Font(name='Arial', bold=True, color=WHITE, size=16)
ws4.cell(row=1, column=1).alignment = Alignment(horizontal='left', vertical='center')

# Flow boxes
flow = [("School\nAdmin / Teacher", BLUE), ("KlassApp\nEngine", GREEN), ("WhatsApp\nAPI", AMBER), ("Parent's Phone", DARK)]
for i, (label, color) in enumerate(flow):
    col = 2 + i
    fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    ws4.merge_cells(start_row=3, start_column=col, end_row=5, end_column=col)
    ws4.cell(row=3, column=col, value=label).font = Font(name='Arial', bold=True, color=WHITE, size=14)
    ws4.cell(row=3, column=col).fill = fill
    ws4.cell(row=3, column=col).alignment = center_align
    ws4.row_dimensions[3].height = 25
    ws4.row_dimensions[4].height = 25
    ws4.row_dimensions[5].height = 25
    if i < 3:
        ws4.cell(row=4, column=col+1, value="→").font = Font(name='Arial', bold=True, color=MID_GRAY, size=18)
        ws4.cell(row=4, column=col+1).alignment = center_align

# Dashboard box
ws4.merge_cells('B7:E7')
ws4.cell(row=7, column=2, value="Delivery Dashboard  ·  Analytics  ·  Failure Escalation  ·  AI Import").font = Font(name='Arial', color=TEXT_DARK, size=11)
ws4.cell(row=7, column=2).fill = light_fill
ws4.cell(row=7, column=2).border = thin_border
ws4.cell(row=7, column=2).alignment = center_align

# Features
feat = [
    "Schools publish grades, fees, or attendance → KlassApp formats and delivers instantly.",
    "Parents reply with keywords (\"grades\", \"fees\", \"absent\") → instant AI response.",
    "Multi-child support: one parent, multiple students, separate notification streams.",
    "AI-powered import: Excel/CSV parsing, smart delivery routing, failure escalation.",
    "LIN integration: connects to Uganda's national Learner Identification Number system.",
]
for i, f in enumerate(feat):
    ws4.merge_cells(f'B{9+i}:F{9+i}')
    ws4.cell(row=9+i, column=2, value=f).font = Font(name='Arial', color=TEXT_DARK, size=11)
    ws4.cell(row=9+i, column=2).alignment = Alignment(horizontal='left', vertical='center')
    ws4.cell(row=9+i, column=2).fill = light_fill if i % 2 == 0 else PatternFill(fill_type=None)

# ═══════════════════════════════════════
# SHEET 5 — DIFFERENTIATORS
# ═══════════════════════════════════════
ws5 = wb.create_sheet("5. Differentiators")
set_col_widths(ws5, [3, 30, 3, 30, 3, 30])

style_range(ws5, 1, 1, 1, 6, fill=header_fill)
ws5.merge_cells('A1:F1')
ws5.cell(row=1, column=1, value="DIFFERENTIATORS — Why KlassApp Wins").font = Font(name='Arial', bold=True, color=WHITE, size=16)
ws5.cell(row=1, column=1).alignment = Alignment(horizontal='left', vertical='center')

cols_data = [
    ("WhatsApp-First", GREEN, [
        "Works inside the app parents already\nuse daily",
        "Zero friction: no download, no\naccount, no password",
        "98% penetration in target market",
        "Zero-rated on MTN & Airtel —\nfree to receive",
    ]),
    ("LIN Integration", BLUE, [
        "Connects to Uganda's national\nstudent ID database",
        "Parents self-register by texting\ntheir child's LIN",
        "Creates a nationwide parent\nnetwork — not just per-school",
        "Moat: no competitor connects to\nthe national LIN system",
    ]),
    ("Platform vs. Point Solution", AMBER, [
        "Layers on top of existing systems\n— no rip-and-replace",
        "Roles: admin, teacher, bursar,\nlibrarian, nurse, secretary",
        "Expanding: counsellor, sports,\ntransport, PTA",
        "Student data follows the child\nacross school transfers",
    ]),
]

for ci, (col_title, col_color, items) in enumerate(cols_data):
    scol = 2 + ci * 2
    fill = PatternFill(start_color=col_color, end_color=col_color, fill_type='solid')
    light = PatternFill(start_color=OFF_WHITE, end_color=OFF_WHITE, fill_type='solid')
    
    ws5.merge_cells(start_row=3, start_column=scol, end_row=3, end_column=scol+1)
    ws5.cell(row=3, column=scol, value=col_title).font = Font(name='Arial', bold=True, color=WHITE, size=14)
    ws5.cell(row=3, column=scol).fill = fill
    ws5.cell(row=3, column=scol).alignment = center_align
    
    for ii, item in enumerate(items):
        r = 5 + ii * 3
        ws5.merge_cells(start_row=r, start_column=scol, end_row=r+1, end_column=scol+1)
        ws5.cell(row=r, column=scol, value=item).font = Font(name='Arial', color=TEXT_DARK, size=11)
        ws5.cell(row=r, column=scol).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws5.cell(row=r, column=scol).fill = light
        ws5.cell(row=r, column=scol).border = thin_border
        ws5.row_dimensions[r].height = 28
        ws5.row_dimensions[r+1].height = 28

# ═══════════════════════════════════════
# SHEET 6 — TRACTION
# ═══════════════════════════════════════
ws6 = wb.create_sheet("6. Traction")
set_col_widths(ws6, [4, 25, 25, 25, 20])

style_range(ws6, 1, 1, 1, 5, fill=header_fill)
ws6.merge_cells('A1:E1')
ws6.cell(row=1, column=1, value="TRACTION — Early Adopters").font = Font(name='Arial', bold=True, color=WHITE, size=16)
ws6.cell(row=1, column=1).alignment = Alignment(horizontal='left', vertical='center')

# Big stat cards
stats = [("5", "schools live in testnet", GREEN), ("150+", "teachers active on the platform", BLUE), ("2,500+", "students on the system", AMBER)]
for i, (num, label, color) in enumerate(stats):
    col = 2 + i
    fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    ws6.merge_cells(start_row=3, start_column=col, end_row=5, end_column=col)
    ws6.cell(row=3, column=col, value=f"{num}\n\n{label}").font = Font(name='Arial', bold=True, color=WHITE, size=36)
    ws6.cell(row=3, column=col).fill = fill
    ws6.cell(row=3, column=col).alignment = center_align
    ws6.row_dimensions[3].height = 50
    ws6.row_dimensions[4].height = 30
    ws6.row_dimensions[5].height = 10

# Product Progress
ws6.merge_cells('B8:C8')
ws6.cell(row=8, column=2, value="PRODUCT PROGRESS").font = Font(name='Arial', bold=True, color=DARK, size=13)

phases = [
    ("Phase 1 — Core", "✓ Complete", GREEN, "Grade publishing, fee reminders, attendance alerts, interactive bot menu"),
    ("Phase 2 — Engage", "✓ Complete", GREEN, "Smart message delivery, notification queue, delivery dashboard, multi-child flow"),
    ("Phase 3 — Scale", "⟳ In Progress", AMBER, "LIN integration, CSV bulk import, parent self-registration, nationwide network"),
]
for i, (label, status, color, desc) in enumerate(phases):
    r = 10 + i * 2
    ws6.cell(row=r, column=2, value=label).font = Font(name='Arial', bold=True, color=DARK, size=11)
    ws6.cell(row=r, column=3, value=status).font = Font(name='Arial', bold=True, color=color, size=11)
    ws6.merge_cells(f'D{r}:E{r}')
    ws6.cell(row=r, column=4, value=desc).font = Font(name='Arial', color=TEXT_GRAY, size=10)
    ws6.cell(row=r, column=4).alignment = left_align
    ws6.row_dimensions[r].height = 22

# Green accent at bottom
for c in range(1, 6):
    ws6.cell(row=18, column=c).fill = green_fill

# ═══════════════════════════════════════
# SHEET 7 — MARKET
# ═══════════════════════════════════════
ws7 = wb.create_sheet("7. Market")
set_col_widths(ws7, [4, 18, 22, 16, 10, 25, 10])

style_range(ws7, 1, 1, 1, 7, fill=header_fill)
ws7.merge_cells('A1:G1')
ws7.cell(row=1, column=1, value="MARKET — Opportunity in Uganda & East Africa").font = Font(name='Arial', bold=True, color=WHITE, size=16)
ws7.cell(row=1, column=1).alignment = Alignment(horizontal='left', vertical='center')

# Market sizing
ws7.merge_cells('B3:D3')
ws7.cell(row=3, column=2, value="MARKET SIZING — Uganda").font = Font(name='Arial', bold=True, color=DARK, size=13)

# Table headers
table_headers = ["", "Description", "Parent Connections"]
for i, h in enumerate(table_headers):
    ws7.cell(row=5, column=2+i, value=h).font = Font(name='Arial', bold=True, color=WHITE, size=11)
    ws7.cell(row=5, column=2+i).fill = PatternFill(start_color=DARK, end_color=DARK, fill_type='solid')
    ws7.cell(row=5, column=2+i).border = thin_border
    ws7.cell(row=5, column=2+i).alignment = center_align

market_data = [
    ("TAM", "74,000 schools × 500 avg parents", "37M connections"),
    ("SAM", "~30,000 schools with smartphone access", "15M connections"),
    ("SOM (Y3)", "~5,000 schools adopted", "2.5M connections"),
]
for i, (label, desc, num) in enumerate(market_data):
    ws7.cell(row=6+i, column=2, value=label).font = Font(name='Arial', bold=True, color=BLUE, size=11)
    ws7.cell(row=6+i, column=3, value=desc).font = Font(name='Arial', color=TEXT_DARK, size=11)
    ws7.cell(row=6+i, column=4, value=num).font = Font(name='Arial', bold=True, color=TEXT_DARK, size=11)
    for c in range(2, 5):
        ws7.cell(row=6+i, column=c).border = thin_border
        ws7.cell(row=6+i, column=c).alignment = left_align if c < 4 else center_align

# Revenue callout
ws7.merge_cells('B10:D10')
ws7.cell(row=10, column=2, value="At SOM: 5,000 schools × $22.30/mo blended = $1.34M ARR").font = Font(name='Arial', bold=True, color=WHITE, size=12)
ws7.cell(row=10, column=2).fill = blue_fill
ws7.cell(row=10, column=2).alignment = center_align

# Growth drivers
ws7.merge_cells('F3:G3')
ws7.cell(row=3, column=6, value="KEY GROWTH DRIVERS").font = Font(name='Arial', bold=True, color=DARK, size=13)

drivers = [
    ("LIN Mandate", "Every student already in a national database —\nwe just connect the parent."),
    ("Smartphone Adoption", "$40 devices now common; WhatsApp is the\nde facto OS for most Ugandans."),
    ("Mobile Money", "70%+ of school fees paid via mobile money —\nnotifications drive action."),
    ("No Direct Competitor", "No WhatsApp-first, LIN-integrated school\nplatform exists in East Africa."),
]
for i, (title, desc) in enumerate(drivers):
    r = 5 + i * 3
    ws7.cell(row=r, column=6).fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type='solid')
    ws7.merge_cells(f'F{r+1}:G{r+1}')
    ws7.cell(row=r+1, column=6, value=title).font = Font(name='Arial', bold=True, color=DARK, size=12)
    ws7.merge_cells(f'F{r+2}:G{r+2}')
    ws7.cell(row=r+2, column=6, value=desc).font = Font(name='Arial', color=TEXT_GRAY, size=10)

# ═══════════════════════════════════════
# SHEET 8 — FINANCIALS
# ═══════════════════════════════════════
ws8 = wb.create_sheet("8. KlassNomics")
set_col_widths(ws8, [4, 30, 16, 16, 16, 16])

style_range(ws8, 1, 1, 1, 6, fill=header_fill)
ws8.merge_cells('A1:F1')
ws8.cell(row=1, column=1, value="KLASSNOMICS — Revenue Model").font = Font(name='Arial', bold=True, color=WHITE, size=16)
ws8.cell(row=1, column=1).alignment = Alignment(horizontal='left', vertical='center')

# Revenue table
fin_headers = ["Metric", "25 Schools", "50 Schools", "100 Schools", "200 Schools"]
for i, h in enumerate(fin_headers):
    ws8.cell(row=3, column=2+i, value=h).font = Font(name='Arial', bold=True, color=WHITE, size=11)
    ws8.cell(row=3, column=2+i).fill = PatternFill(start_color=DARK, end_color=DARK, fill_type='solid')
    ws8.cell(row=3, column=2+i).border = thin_border
    ws8.cell(row=3, column=2+i).alignment = center_align

fin_data = [
    ("Gross Revenue (USD/yr)", ["$6,689", "$13,378", "$26,757", "$53,514"]),
    ("Infra & AI Costs (~12%)", ["$803", "$1,605", "$3,211", "$6,422"]),
    ("Net Revenue (USD/yr)", ["$5,886", "$11,773", "$23,546", "$47,092"]),
    ("Gross Margin", ["88%", "88%", "88%", "88%"]),
]
for ri, (label, vals) in enumerate(fin_data):
    r = 4 + ri
    is_net = ri == 2
    ws8.cell(row=r, column=2, value=label).font = Font(name='Arial', bold=is_net, color=BLUE if is_net else TEXT_DARK, size=11)
    ws8.cell(row=r, column=2).fill = accent_blue_fill if is_net else PatternFill(fill_type=None)
    for vi, v in enumerate(vals):
        ws8.cell(row=r, column=3+vi, value=v).font = Font(name='Arial', bold=is_net, color=TEXT_DARK, size=11)
        ws8.cell(row=r, column=3+vi).fill = accent_blue_fill if is_net else PatternFill(fill_type=None)
        ws8.cell(row=r, column=3+vi).alignment = right_align
    for c in range(2, 7):
        ws8.cell(row=r, column=c).border = thin_border

# Pricing tiers
ws8.merge_cells('B9:C9')
ws8.cell(row=9, column=2, value="PRICING TIERS").font = Font(name='Arial', bold=True, color=DARK, size=13)

tiers = [
    ("Basic", "UGX 50K/mo", "$13.50", "Manual entry, CSV import, attendance, parent portal", BLUE),
    ("Pro", "UGX 100K/mo", "$27.00", "AI Excel import, WhatsApp alerts, analytics, push notifications", GREEN),
    ("Enterprise", "UGX 150K/mo", "$40.50", "Register scanning, full AI suite, custom branding, SSO", AMBER),
]
for i, (name, price, usd, desc, color) in enumerate(tiers):
    col = 2 + i
    fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    ws8.merge_cells(start_row=11, start_column=col, end_row=12, end_column=col)
    ws8.cell(row=11, column=col, value=name).font = Font(name='Arial', bold=True, color=WHITE, size=14)
    ws8.cell(row=11, column=col).fill = fill
    ws8.cell(row=11, column=col).alignment = center_align
    ws8.merge_cells(start_row=13, start_column=col, end_row=13, end_column=col)
    ws8.cell(row=13, column=col, value=f"{price}\n({usd})").font = Font(name='Arial', bold=True, color=TEXT_DARK, size=11)
    ws8.cell(row=13, column=col).alignment = center_align
    ws8.cell(row=13, column=col).fill = light_fill
    ws8.merge_cells(start_row=14, start_column=col, end_row=14, end_column=col)
    ws8.cell(row=14, column=col, value=desc).font = Font(name='Arial', color=TEXT_GRAY, size=9)
    ws8.cell(row=14, column=col).alignment = center_align
    ws8.cell(row=14, column=col).fill = light_fill

# ═══════════════════════════════════════
# SHEET 9 — GO-TO-MARKET
# ═══════════════════════════════════════
ws9 = wb.create_sheet("9. Go-to-Market")
set_col_widths(ws9, [4, 40, 30, 20, 16])

style_range(ws9, 1, 1, 1, 5, fill=header_fill)
ws9.merge_cells('A1:E1')
ws9.cell(row=1, column=1, value="GO-TO-MARKET — How We Grow").font = Font(name='Arial', bold=True, color=WHITE, size=16)
ws9.cell(row=1, column=1).alignment = Alignment(horizontal='left', vertical='center')

gtm = [
    ("Land", "Free Starter tier for small schools (up to 200 students). Zero-risk trial."),
    ("Expand", "Schools upgrade to Pro/Enterprise as they grow. AI, analytics, branding."),
    ("Network Effects", "Every parent onboarded is a referral. Multi-school parents create pull demand."),
    ("LIN Scale", "Government mandate + national ID = instant parent connection at scale."),
    ("Partnerships", "Ministry of Education, school associations, education NGOs, mobile money."),
    ("Low CAC", "Word-of-mouth via teachers and parents. WhatsApp groups = organic distribution."),
]
for i, (label, desc) in enumerate(gtm):
    r = 3 + i
    ws9.cell(row=r, column=2, value=label).font = Font(name='Arial', bold=True, color=BLUE, size=11)
    ws9.merge_cells(f'C{r}:D{r}')
    ws9.cell(row=r, column=3, value=desc).font = Font(name='Arial', color=TEXT_DARK, size=11)
    ws9.cell(row=r, column=3).alignment = left_align

# Flywheel
ws9.merge_cells('E3:E4')
ws9.cell(row=3, column=5, value="GROWTH FLYWHEEL").font = Font(name='Arial', bold=True, color=DARK, size=11)
ws9.cell(row=3, column=5).alignment = center_align

flywheel = ["Free schools onboard", "Parents engaged", "Parents become advocates", "More schools join"]
for i, step in enumerate(flywheel):
    ws9.cell(row=5+i, column=5, value=f"{i+1}. {step}").font = Font(name='Arial', color=TEXT_DARK, size=10)
    ws9.cell(row=5+i, column=5).fill = light_fill

# Target box
ws9.merge_cells('B13:D13')
ws9.cell(row=13, column=2, value="18-MONTH TARGET").font = Font(name='Arial', bold=True, color=WHITE, size=12)
ws9.cell(row=13, column=2).fill = PatternFill(start_color=DARK, end_color=DARK, fill_type='solid')
ws9.cell(row=13, column=2).alignment = center_align
ws9.merge_cells('B14:D14')
ws9.cell(row=14, column=2, value="200 schools  →  $53K ARR  →  $47K net  →  Path to Series A").font = Font(name='Arial', bold=True, color=GREEN, size=12)
ws9.cell(row=14, column=2).fill = light_fill
ws9.cell(row=14, column=2).border = thin_border
ws9.cell(row=14, column=2).alignment = center_align

# ═══════════════════════════════════════
# SHEET 10 — TEAM & ASK
# ═══════════════════════════════════════
ws10 = wb.create_sheet("10. Team & Ask")
set_col_widths(ws10, [4, 35, 30, 4, 35])

style_range(ws10, 1, 1, 1, 5, fill=header_fill)
ws10.merge_cells('A1:E1')
ws10.cell(row=1, column=1, value="TEAM & THE ASK").font = Font(name='Arial', bold=True, color=WHITE, size=16)
ws10.cell(row=1, column=1).alignment = Alignment(horizontal='left', vertical='center')

# Team
ws10.merge_cells('B3:C3')
ws10.cell(row=3, column=2, value="TEAM").font = Font(name='Arial', bold=True, color=DARK, size=14)

ws10.merge_cells('B5:C7')
ws10.cell(row=5, column=2, value="Elijah (Rasta) — Founder & CEO\n\nFull-stack engineer. Built KlassApp from scratch. Deep understanding of Uganda's education ecosystem.").font = Font(name='Arial', color=TEXT_DARK, size=11)
ws10.cell(row=5, column=2).fill = light_fill
ws10.cell(row=5, column=2).border = thin_border
ws10.cell(row=5, column=2).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
ws10.row_dimensions[5].height = 50
ws10.row_dimensions[6].height = 20
ws10.row_dimensions[7].height = 20

ws10.merge_cells('B9:C10')
ws10.cell(row=9, column=2, value="Co-founder — Technical & Operations Lead\n\nTechnical architecture and operational leadership.").font = Font(name='Arial', color=TEXT_DARK, size=11)
ws10.cell(row=9, column=2).fill = light_fill
ws10.cell(row=9, column=2).border = thin_border
ws10.cell(row=9, column=2).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
ws10.row_dimensions[9].height = 40
ws10.row_dimensions[10].height = 20

# The Ask
ws10.merge_cells('E3:E3')
ws10.cell(row=3, column=5, value="THE ASK").font = Font(name='Arial', bold=True, color=DARK, size=14)

ws10.merge_cells('E5:E5')
ws10.cell(row=5, column=5, value="Raising a Pre-Seed Round").font = Font(name='Arial', bold=True, color=WHITE, size=13)
ws10.cell(row=5, column=5).fill = PatternFill(start_color=DARK, end_color=DARK, fill_type='solid')
ws10.cell(row=5, column=5).alignment = center_align

ws10.merge_cells('E6:E6')
ws10.cell(row=6, column=5, value="USE OF FUNDS").font = Font(name='Arial', bold=True, color=DARK, size=11)

alloc = [
    ("40%", "Engineering & AI", GREEN),
    ("30%", "Sales & School Onboarding", BLUE),
    ("20%", "Infrastructure & WhatsApp API", AMBER),
    ("10%", "Legal, Compliance & Operations", MID_GRAY),
]
for i, (pct, desc, color) in enumerate(alloc):
    r = 7 + i
    ws10.cell(row=r, column=5, value=desc).font = Font(name='Arial', color=TEXT_DARK, size=11)
    ws10.cell(row=r, column=5).alignment = left_align

# Target bar
ws10.merge_cells('B13:E13')
ws10.cell(row=13, column=2, value="Target: 200 schools in 18 months  →  $53K ARR  →  $47K net  →  Path to Series A").font = Font(name='Arial', bold=True, color=GREEN, size=13)
ws10.cell(row=13, column=2).fill = light_fill
ws10.cell(row=13, column=2).border = thin_border
ws10.cell(row=13, column=2).alignment = center_align

# ═══════════════════════════════════════
# SHEET 11 — CLOSING
# ═══════════════════════════════════════
ws11 = wb.create_sheet("11. Closing")
set_col_widths(ws11, [20, 40, 20, 20])

# Dark header
for c in range(1, 9):
    ws11.cell(row=1, column=c).fill = PatternFill(start_color=DARK, end_color=DARK, fill_type='solid')

ws11.merge_cells('B4:F4')
ws11.cell(row=4, column=2, value="Thank You").font = Font(name='Arial', bold=True, color=BLUE, size=42)
ws11.cell(row=4, column=2).alignment = Alignment(horizontal='center', vertical='center')

# Green accent
for c in range(4, 7):
    ws11.cell(row=6, column=c).fill = green_fill

ws11.merge_cells('B8:F8')
ws11.cell(row=8, column=2, value="Let's put a school in every parent's pocket.").font = Font(name='Arial', color=TEXT_GRAY, size=18)
ws11.cell(row=8, column=2).alignment = Alignment(horizontal='center', vertical='center')

ws11.merge_cells('B10:F10')
ws11.cell(row=10, column=2, value="KlassApp — Smarter schools start here.").font = Font(name='Arial', color=TEXT_DARK, size=13)
ws11.cell(row=10, column=2).alignment = Alignment(horizontal='center', vertical='center')

ws11.merge_cells('B12:F12')
ws11.cell(row=12, column=2, value="ugasch.com  ·  WhatsApp: +256 767 538805").font = Font(name='Arial', color=TEXT_GRAY, size=11)
ws11.cell(row=12, column=2).alignment = Alignment(horizontal='center', vertical='center')

ws11.merge_cells('B15:F15')
ws11.cell(row=15, column=2, value="Confidential — June 2026").font = Font(name='Arial', color=MID_GRAY, size=10)
ws11.cell(row=15, column=2).alignment = Alignment(horizontal='center', vertical='center')

# ═══════════════════════════════════════
# SAVE
# ═══════════════════════════════════════
wb.save("/Users/mac/projects/KlassApp/docs/KlassApp_Pitch_Deck.xlsx")
print("Saved to /Users/mac/projects/KlassApp/docs/KlassApp_Pitch_Deck.xlsx")
